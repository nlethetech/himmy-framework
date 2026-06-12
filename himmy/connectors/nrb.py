"""Nepal connectors: Nepal Rastra Bank (NRB) — forex + macroeconomic reports.

Fetches DIRECTLY from NRB's public surface (no intermediary):

* :meth:`NRBClient.forex` / :meth:`latest_forex` — the public forex JSON API
  (``/api/forex/v1/rates``), normalized into :class:`ForexRate`s.
* :meth:`list_macro_reports` — the monthly "Current Macroeconomic & Financial
  Situation" reports, listed via NRB's category RSS feed (Nepali / English /
  Tables variants, with language + data period parsed out).
* :meth:`fetch_latest_macro_workbook` / :meth:`fetch_macro_workbook` /
  :meth:`parse_workbook` — auto-discover and download the macro Excel. NRB's
  'Tables' report URLs serve the ``.xlsx`` directly (a download response), so
  ``fetch_macro_workbook`` parses the fetched bytes as a workbook (falling back to
  scanning an HTML page for a linked spreadsheet), and ``fetch_latest_macro_workbook``
  needs no URL at all — it finds the newest Tables report and parses every sheet
  with ``openpyxl``. (Live: ~93 sheets of CPI/WPI/GDP/etc. data.)
"""

from __future__ import annotations

import io
import re
import zipfile
from datetime import date

from pydantic import BaseModel, Field, ValidationError, field_validator

from himmy.connectors._feeds import entry_cap, parse_feed
from himmy.connectors.fetcher import Fetcher, HttpxFetcher, get_json
from himmy.connectors.models import ForexRate, MacroReport, Workbook
from himmy.core.errors import HimmyError
from himmy.toolkit._net import build_pinned_transport, guard_url

NRB_FOREX_API = "https://www.nrb.org.np/api/forex/v1/rates"
NRB_MACRO_FEED = "https://www.nrb.org.np/category/current-macroeconomic-situation/feed/"

# Ceiling on a workbook's TOTAL decompressed size. The download is already capped
# (HttpxFetcher.max_bytes, ~8MB), but an .xlsx is a zip: a small compressed file can
# carry a huge shared-strings table / sheet that openpyxl would inflate into memory
# (a "zip bomb"). We reject before openpyxl ever sees it. Live NRB Tables workbooks
# (~93 sheets of CPI/WPI/GDP data) decompress well under this bound.
_MAX_WORKBOOK_UNCOMPRESSED_BYTES = 256 * 1024 * 1024

# A link to a downloadable spreadsheet on the NRB site.
_XLS_RE = re.compile(r'https?://[^\s"\'<>]+?\.(?:xlsx|xls)', re.IGNORECASE)
_PERIOD_RE = re.compile(
    r"(annual|first|two|three|four|five|six|seven|eight|nine|ten|eleven)"
    r"[ -]?months?[^0-9]*([0-9]{4}[\.\-/][0-9]{2,4})",
    re.IGNORECASE,
)


def _to_float(value: object) -> float | None:
    try:
        return float(value)  # type: ignore[arg-type]
    except (TypeError, ValueError):
        return None


def _detect_language(title: str) -> str:
    low = title.lower()
    if "tables" in low:
        return "tables"
    if "english" in low:
        return "english"
    if "nepali" in low:
        return "nepali"
    return "unknown"


def _detect_period(title: str) -> str | None:
    match = _PERIOD_RE.search(title)
    if match:
        return f"{match.group(1).lower()}-months {match.group(2)}"
    return None


def _find_workbook_link(html: str) -> str | None:
    match = _XLS_RE.search(html)
    return match.group(0) if match else None


def _looks_like_workbook(data: bytes) -> bool:
    """True when ``data`` is a spreadsheet by magic bytes (xlsx zip / legacy xls)."""
    return data[:4] == b"PK\x03\x04" or data[:4] == b"\xd0\xcf\x11\xe0"


def _guard_decompressed_size(data: bytes, limit: int | None = None) -> None:
    """Reject a zip-bomb ``.xlsx`` before openpyxl inflates it into memory.

    An ``.xlsx`` is a zip container, so its compressed size (already capped on
    download) says nothing about how much memory parsing it will cost: a tiny file
    can declare a multi-GB shared-strings table or sheet. The central directory
    records each member's uncompressed ``file_size``; we sum those (and check each
    member) WITHOUT decompressing anything, and raise if the total — or any single
    member — would exceed ``limit`` (defaulting to the module cap, read at call time
    so it stays configurable/patchable). Non-zip inputs (legacy ``.xls``) are left
    for openpyxl/the caller to handle.
    """
    if limit is None:
        limit = _MAX_WORKBOOK_UNCOMPRESSED_BYTES
    if data[:4] != b"PK\x03\x04":  # only the zip-backed format can be a zip bomb
        return
    try:
        with zipfile.ZipFile(io.BytesIO(data)) as archive:
            total = 0
            for info in archive.infolist():
                if info.file_size > limit:
                    raise HimmyError(
                        f"workbook member {info.filename!r} declares "
                        f"{info.file_size} uncompressed bytes, over the "
                        f"{limit}-byte safety cap (possible zip bomb)"
                    )
                total += info.file_size
                if total > limit:
                    raise HimmyError(
                        f"workbook decompresses to over {limit} bytes "
                        f"(possible zip bomb); refusing to parse"
                    )
    except zipfile.BadZipFile:
        return  # not a real zip — let the parser surface the real error


# ----------------------------------------------------------- forex payload schema
# NRB's forex API shape, validated with pydantic so structure drift fails LOUDLY
# (a clear HimmyError) instead of as KeyErrors or silently empty results. Value-
# level noise stays tolerated (an unparseable buy/sell becomes None, a missing
# unit becomes 1) — only the *structure* is load-bearing.


class _ForexCurrency(BaseModel):
    """The ``currency`` object inside one NRB forex rate entry."""

    iso3: str = ""
    name: str = ""
    unit: int = 1

    @field_validator("unit", mode="before")
    @classmethod
    def _unit_default(cls, value: object) -> object:
        return value or 1


class _ForexEntry(BaseModel):
    """One per-currency rate entry inside a day's ``rates`` list."""

    currency: _ForexCurrency = Field(default_factory=_ForexCurrency)
    buy: float | None = None
    sell: float | None = None

    @field_validator("currency", mode="before")
    @classmethod
    def _currency_default(cls, value: object) -> object:
        return value or {}

    @field_validator("buy", "sell", mode="before")
    @classmethod
    def _price(cls, value: object) -> float | None:
        return _to_float(value)


class _ForexDay(BaseModel):
    """One published day: a date plus its per-currency rates."""

    date: str = ""
    rates: list[_ForexEntry] = Field(default_factory=list)

    @field_validator("rates", mode="before")
    @classmethod
    def _rates_default(cls, value: object) -> object:
        return value if value is not None else []


class _ForexData(BaseModel):
    """The ``data`` envelope: ``payload`` is the list of days."""

    payload: list[_ForexDay]


class _ForexResponse(BaseModel):
    """The top-level NRB forex API response (``{"data": {"payload": [...]}}``)."""

    data: _ForexData


class NRBClient:
    """A direct client for NRB forex + macroeconomic-situation reports."""

    def __init__(self, fetcher: Fetcher | None = None) -> None:
        """Wire the HTTP fetcher (defaults to the live ``HttpxFetcher``).

        The default fetcher carries the SSRF guard so redirect hops AND the
        spreadsheet links discovered on a (possibly model-supplied) report page
        must resolve to public addresses — never loopback / RFC1918 / IMDS. Its
        transport pins the vetted IP at connect time so a low-TTL name can't rebind
        between the guard's DNS lookup and httpx's own (the rebinding TOCTOU gap).
        """
        self._fetcher = fetcher or HttpxFetcher(
            guard=guard_url, transport=build_pinned_transport()
        )

    # ------------------------------------------------------------------- forex
    def forex(
        self, from_date: str, to_date: str | None = None, *, per_page: int = 100
    ) -> list[ForexRate]:
        """Return per-currency buy/sell rates for a date (or ``from``..``to`` range).

        Dates are ISO ``YYYY-MM-DD``. Each currency's ``unit`` is the quantity the
        buy/sell price is quoted for (e.g. INR is quoted per 100). NRB caps
        ``per_page`` (the number of days per page) at 100; larger values return
        nothing, so the default and effective maximum is 100. The response is
        schema-validated: if NRB changes the payload structure this raises a clear
        :class:`HimmyError` rather than mis-parsing or returning silent emptiness.
        """
        per_page = min(per_page, 100)
        to_date = to_date or from_date
        url = (
            f"{NRB_FOREX_API}?from={from_date}&to={to_date}&per_page={per_page}&page=1"
        )
        data = get_json(self._fetcher, url)
        try:
            parsed = _ForexResponse.model_validate(data)
        except ValidationError as exc:
            raise HimmyError(
                f"NRB forex response no longer matches the expected "
                f"{{data: {{payload: [...]}}}} shape (schema drift?): {exc}"
            ) from exc
        return [
            ForexRate(
                date=day.date,
                currency_iso3=entry.currency.iso3,
                currency_name=entry.currency.name,
                unit=entry.currency.unit,
                buy=entry.buy,
                sell=entry.sell,
            )
            for day in parsed.data.payload
            for entry in day.rates
        ]

    def latest_forex(self) -> list[ForexRate]:
        """Convenience: today's published rates."""
        return self.forex(date.today().isoformat())

    # ------------------------------------------------------ macroeconomic data
    def list_macro_reports(self, *, limit: int = 20) -> list[MacroReport]:
        """List the latest 'Current Macroeconomic & Financial Situation' reports.

        Sourced from NRB's category RSS feed (the HTML listing is JS-rendered).
        Each report comes in Nepali / English / Tables variants — the ``language``
        field distinguishes them, and ``period`` parses the data window. An
        unparseable feed (error page, format drift) raises a clear
        :class:`HimmyError` instead of returning a silent empty list.
        """
        raw = self._fetcher.get_bytes(NRB_MACRO_FEED)
        feed = parse_feed(raw, source="nrb-macro")
        reports: list[MacroReport] = []
        for entry in feed.entries[: entry_cap(limit)]:
            title = entry.get("title", "")
            reports.append(
                MacroReport(
                    title=title,
                    url=entry.get("link", ""),
                    published=entry.get("published") or entry.get("updated"),
                    language=_detect_language(title),
                    period=_detect_period(title),
                )
            )
        return reports

    def fetch_macro_workbook(self, report_url: str) -> Workbook | None:
        """Download + parse a report's Excel workbook (auto-discovering the source).

        NRB's 'Tables' report URLs serve the ``.xlsx`` *directly* (a download
        response), so this first tries to parse the fetched bytes AS a workbook; if
        it instead got an HTML page, it scans that page for a linked ``.xlsx``/
        ``.xls`` and fetches that. Returns None only when neither yields a workbook.
        """
        data = self._fetcher.get_bytes(report_url)
        if _looks_like_workbook(data):
            try:
                workbook = self.parse_workbook(data)
                workbook.source_url = report_url
                return workbook
            except Exception:  # noqa: BLE001 - not really a workbook; try the HTML path
                pass
        link = _find_workbook_link(data.decode("utf-8", "replace"))
        if link is None:
            return None
        workbook = self.parse_workbook(self._fetcher.get_bytes(link))
        workbook.source_url = link
        return workbook

    def fetch_latest_macro_workbook(self) -> Workbook | None:
        """Auto-discover + fetch the latest macro **Tables** workbook (no URL needed).

        Lists the macro reports, picks the newest 'Tables' variant (the one that
        carries the spreadsheet), and downloads + parses it.
        """
        for report in self.list_macro_reports(limit=20):
            if report.language == "tables":
                return self.fetch_macro_workbook(report.url)
        return None

    @staticmethod
    def parse_workbook(data: bytes) -> Workbook:
        """Parse ``.xlsx`` bytes into ``{sheet name: rows}`` (every cell, as-is)."""
        try:
            import openpyxl
        except ImportError as exc:  # pragma: no cover - only without the extra
            raise HimmyError(
                "Excel parsing requires the [connectors] extra "
                "(pip install 'himmy[connectors]')."
            ) from exc
        import warnings

        # Bound the DECOMPRESSED size before openpyxl inflates the (model/attacker-
        # influenced) bytes — the download cap only bounds the compressed download,
        # so a small zip-bomb .xlsx could otherwise exhaust memory at parse time.
        _guard_decompressed_size(data)

        with warnings.catch_warnings():
            # NRB workbooks use extensions openpyxl warns (harmlessly) about.
            warnings.simplefilter("ignore")
            book = openpyxl.load_workbook(
                io.BytesIO(data), read_only=True, data_only=True
            )
            sheets: dict[str, list[list[object]]] = {}
            for sheet in book.worksheets:
                sheets[sheet.title] = [
                    list(row) for row in sheet.iter_rows(values_only=True)
                ]
            book.close()
        return Workbook(sheets=sheets)


__all__ = ["NRBClient", "NRB_FOREX_API", "NRB_MACRO_FEED"]
